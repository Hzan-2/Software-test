import os

import openpyxl
import warnings

from config.Config import DOWNLOAD_FILE


def jugde_Excel_color(rows,column,Sheet='Sheet0'):
    '''
    :param rows: 行
    :param column: 列
    :param Sheet: 表单名称
    :return: 布尔值
    '''
    warnings.filterwarnings('ignore')
    for fileName in os.listdir(os.path.join(DOWNLOAD_FILE)):
        if '.xlsx' in str(fileName) or '.xls' in str(fileName):
            workbook = openpyxl.load_workbook(os.path.join(DOWNLOAD_FILE,fileName))
            worksheet = workbook[Sheet]
            is_color = worksheet.cell(row=int(rows),column=int(column)).fill.start_color.rgb
            if is_color == '00000000':
                return False
            else:
                return True


def get_Excel_Text(rows,column,Sheet='Sheet0'):
    '''
    :param rows: 行
    :param column: 列
    :param Sheet: 表单名称
    :return: 文本值
    '''
    warnings.filterwarnings('ignore')
    for fileName in os.listdir(os.path.join(DOWNLOAD_FILE)):
        if '.xlsx' in str(fileName) or '.xls' in str(fileName):
            workbook = openpyxl.load_workbook(os.path.join(DOWNLOAD_FILE, fileName))
            worksheet = workbook[Sheet]
            return worksheet.cell(row=int(rows),column=int(column)).value